import matplotlib.pyplot as plt
import openpyxl
import requests
from bs4 import BeautifulSoup
from pandas import DataFrame

url = 'https://www.sportspress.cn/database/integral_2022_378/'
header = {"User-Agent":
              "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36 Edg/114.0.1823.51"}
r = requests.get(url, headers=header)
r.encoding = "utf-8"
r.text
soup = BeautifulSoup(r.text, 'html.parser')
# print(soup)数据获取（后续不输出）
# 将网页转化为HTML文本

zhongchao = open("HTML文本.txt", "w", encoding="utf8")
zhongchao.write(r.text)  # 写入内容
zhongchao.close()

zhongchao_a = soup.find_all("div", attrs={"class": "row1 row2 color_1"})
zhongchao_b = []
for i in range(len(zhongchao_a)):
    zhongchao_b.append(zhongchao_a[i].text)
a = []
for i in range(18):
    j = zhongchao_b[i].split()
    a.append(j)
paiming = []
duiming = []
shengchang = []
defeng = []
shenglv = []
for i in range(10):
    paiming.append(a[i][0])  # 将排名保存到paiming列表中
    duiming.append(a[i][1])  # 将队名保存到duiming列表中
    shengchang.append(a[i][3])  # 将胜场数保存到shengchang列表中
    defeng.append(a[i][6])  # 将得分保存到defeng列表中
    shenglv.append(a[i][9])  # 将胜率保存到shenglv列表中
paiming.insert(0, '队伍排名')
shengchang0 = []
for i in shengchang:
    shengchang0.append(int(i))
# 将shengchang列表中的字符串转换成整数型，并保存到shengchang0中

shengchang1 = []
for i in range(len(shengchang)):
    shengchang1.append(shengchang[i])
duiming0 = []
for i in range(len(duiming)):
    duiming0.append(duiming[i])
defeng0 = []
for i in defeng:
    defeng0.append(int(i))
# 将defeng列表中的字符串转换成整数型，并保存到defeng0中

shenglv0 = []
df = DataFrame({
    'p_str': shenglv
})
shenglv0 = list(df['p_str'].str.strip("%").astype(float) / 100)
shenglv1 = []
for i in shenglv0:
    shenglv1.append(round(i, 2))
# 将百分比胜率转换成

duiming.insert(0, '队伍名称')
shengchang.insert(0, '胜场数')
defeng.insert(0, '得分')
shenglv.insert(0, '胜率')

exfile = open("result.xlsx", "w", encoding="utf8")
exfile.close()


# 创建result.xlsx文件
def write_excel():
    work_book = openpyxl.Workbook()
    sheet = work_book.active
    for i in range(len(paiming)):
        sheet.cell(i + 1, 1, paiming[i])
        sheet.cell(i + 1, 2, duiming[i])
        sheet.cell(i + 1, 3, shengchang[i])
        sheet.cell(i + 1, 4, defeng[i])
        sheet.cell(i + 1, 5, shenglv[i])
    work_book.save('result.xlsx')

write_excel()
# 将数据写入excel表格


plt.rcParams['font.sans-serif'] = ['SimHei']
plt.bar(duiming0, defeng0)
plt.ylim(40, 100)
plt.xticks(rotation=30)
plt.xlabel('队伍名字')
plt.ylabel("得分")
plt.title('排名前十队伍及其得分')
plt.show()
# 绘制排名前十队伍名字及其得分的柱状图



plt.plot(duiming0, shengchang0)
plt.ylim(10, 30)
plt.xticks(rotation=30)
plt.xlabel('队伍名字')
plt.ylabel("胜场数")
plt.title('排名前十队伍及其胜场数')
plt.show()
# 绘制排名前十队伍名字及其胜场数的折线图


plt.scatter(duiming0, shenglv1)
plt.ylim(0, 1.0)
plt.xticks(rotation=30)
plt.xlabel('队伍名字')
plt.ylabel("胜率")
plt.title('排名前十队伍及其胜率')
plt.show()
##绘制排名前十队伍名字及其胜率的散点图
